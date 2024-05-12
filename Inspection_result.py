import streamlit as st
import pandas as pd
import numpy as np
import warnings
warnings.filterwarnings('ignore')
import base64
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import gc
import time
def app():


    # List of image paths (replace with your own image paths)
    image_paths = [
        r"D:\Streamlit\img-4.png"
    ]

    # Create a placeholder for the image
    image_placeholder = st.empty()

    # Display each image in the list for 5 seconds
    for image_path in image_paths:
        # Load the image
        image = Image.open(image_path)
        
        # Display the image in the placeholder
        image_placeholder.image(image, use_column_width=True)
        
        # Wait for 5 seconds before displaying the next image
        time.sleep(0)

    st.markdown("<h1 style='text-align: center; color: blue;'>RFI Inspection Result</h1>", unsafe_allow_html=True)
    st.header('Upload your dataset')

    # Use st.file_uploader with Markdown formatting for the label
    uploaded_file_ccc = st.file_uploader("**_Upload CCC System File_**", type=["xlsx"])

    # Upload SCDB System File with formatted label
    uploaded_file_scdb = st.file_uploader("**_Upload SCDB System File_**", type=["xlsx"])

    # Upload Name Mapping File with formatted label
    uploaded_file_name = st.file_uploader("**_Upload Name Mapping File_**", type=["xlsx"])

    def preprocess_column(column, **kwargs):
        """Preprocess a column using a mapping dictionary for replacements"""
        mapping_dict = kwargs.get('mapping')
        return column.apply(lambda x: mapping_dict.get("".join(str(x).lower().split()), "No Name Match") if pd.notnull(x) else x)

    # Function to preprocess df_ccc DataFrame
    def preprocess_df_ccc(uploaded_file_ccc, mapping_file):
        # Check if files are uploaded
        if uploaded_file_ccc is None or uploaded_file_name is None:
            st.write("Please upload CCC System File and Name Mapping File.")
            return None

        # read excel file as data frame
        df_ccc = pd.read_excel(uploaded_file_ccc)
        mapping_df = pd.read_excel(mapping_file)
        rejected_count = df_ccc[df_ccc['Inspection Result'] == 'Rejected'].shape[0]
        st.write(f"Number of rows with Inspection Result/Priority (Name) as 'Rejected': {rejected_count}")
        other_result_count = df_ccc[~df_ccc['Inspection Result'].isin(['Approved', 'Rejected', 'Cancelled'])].shape[0]
        st.write(f"Number of rows with Inspection Result/Priority (Name) other than 'Accepted', 'Rejected', or 'Cancelled': {other_result_count}")

        #st.write("Task-1")
        # Task - 1 : Rename columns
        df_ccc = df_ccc[['Comm ID', 'Communication Type', 'Company PIC Name:', 
                    'Actual Inspection Finish Date:', 'Workflow - Verified Date', 
                    'Workflow - Verified By', 'Workflow - Accepted Date', 
                    'Workflow - Accepted By','Workflow - Closed Date' , 
                    'Workflow - Closed By', 'Task Name', 
                    'Witness / Review (Type of Inspection)', 'Inspection Result', 'Discipline','Company Intervention Type:']]
        # Convert date columns to datetime and format
        date_cols = ['Actual Inspection Finish Date:', 'Workflow - Verified Date', 
                    'Workflow - Accepted Date', 'Workflow - Closed Date']
        df_ccc.loc[:, date_cols] = df_ccc[date_cols].apply(pd.to_datetime, errors='coerce')
        df_ccc.loc[:, date_cols] = df_ccc[date_cols].apply(lambda x: x.dt.strftime('%d-%b-%y'))
        #st.write("Task-2")
        # Task -2: Map Discipline values
        discipline_mapping = {
            'AB': 'BLD', 'CVL': 'CVL', 'ELE': 'ELE', 'IN': 'CSE', 'IS': 'INS',
            'MEC': 'MEC', 'PA': 'PNT', 'PIP': 'PIP', 'STR': 'STR', 'TC': 'TEL',
            'HV': 'MEC', 'RF': 'REF'
        }
        df_ccc.loc[:, 'Discipline'] = df_ccc['Discipline'].map(discipline_mapping)
        #st.write("Task-3")
        # Task -3a: Update 'Inspection Result' column  "Approved"-----> "Accepted"
        df_ccc.loc[df_ccc['Inspection Result'] == "Approved", 'Inspection Result'] = "Accepted"
        
        # Task -3b: Fill value in 'df[workflow - Accepted BY]' column  based condition
        condition = (df_ccc['Inspection Result'] == 'Cancelled') & (df_ccc['Workflow - Accepted By'].isnull())
        Discipline_mapping = {
            'BLD': 'Kanikodan, Prakashan',
            'CVL': 'Kanikodan, Prakashan',
            'ELE': 'Khurshid, Karawan',
            'CSE': 'Khurshid, Karawan',
            'TEL': 'Khurshid, Karawan',
            'PNT': 'Anjum, Mohammad Muzamil',
            'REF': 'Anjum, Mohammad Muzamil',
            'INS': 'Anjum, Mohammad Muzamil',
            'PIP': 'Ali, Ahsan',
            'STR': 'Agbay, Cecilio Masagpag',
            'MEC': 'Brosoto, Joelito'
        }
        df_ccc.loc[condition, 'Workflow - Accepted By'] = df_ccc.loc[condition, 'Discipline'].map(Discipline_mapping)
        #st.write("Task-4")
        #Task -4: Fill value in 'df[workflow - Accepted Date]' column  based condition
        condition = (df_ccc['Inspection Result'] == 'Cancelled') & (df_ccc['Workflow - Accepted Date'].isnull()) & (df_ccc['Workflow - Accepted By'].notnull() ) 
        df_ccc.loc[condition, 'Workflow - Accepted Date'] = df_ccc.loc[condition, 'Workflow - Verified Date']   
        #st.write("Task-5")
        #Task -5a: Fill value in 'df[Workflow - Closed Date]' column  based condition
        # Define the condition
        condition = (df_ccc['Inspection Result'] == 'Cancelled') & (df_ccc['Workflow - Closed Date'].isnull()) 
        # Fill the values based on the condition
        df_ccc.loc[condition,'Workflow - Closed Date']=df_ccc.loc[condition,'Workflow - Accepted Date']
        #Task -5b: Fill value in 'df[Workflow - Closed By]' column  based condition
        # Define the condition
        condition = (df_ccc['Inspection Result'] == 'Cancelled') & (df_ccc['Workflow - Closed By'].isnull()) 
        # Fill the values based on the condition
        df_ccc.loc[condition,'Workflow - Closed By']=df_ccc.loc[condition,'Workflow - Accepted By']
        #st.write("Task-6")
        # Task -6: Update three columns based on condition
        condition = ((df_ccc['Company Intervention Type:'].isin(['S','R']))) & (df_ccc['Workflow - Closed By'].isnull())

        # Update values based on the condition
        df_ccc['Workflow - Closed Date'] = np.where(condition, df_ccc['Workflow - Accepted Date'], df_ccc['Workflow - Closed Date'])
        df_ccc['Workflow - Closed By'] = np.where(condition, df_ccc['Workflow - Accepted By'], df_ccc['Workflow - Closed By']) 

        condition = ((df_ccc['Company Intervention Type:'].isin(['S','R'])))
        print(condition.sum())
        df_ccc['Company PIC Name:'] = np.where(condition, 'NA', df_ccc['Company PIC Name:'])
        df_ccc['Company PIC Name:'].isnull().sum()
        #st.write("Task-7")
        # Task -7: Replace values in 'Witness / Review (Type of Inspection)'
        df_ccc['Witness / Review (Type of Inspection)'] = df_ccc['Witness / Review (Type of Inspection)'].replace({'Witness': 1, 'Review': 2})

        #st.write("Task-8")
        # Task -8: Name Change Mapping
        mapping_dict = dict(zip(mapping_df.iloc[:, 0].str.lower().str.replace(' ', ''), mapping_df.iloc[:, 1]))
        # Columns to Rename
        columns_to_preprocess = ['Workflow - Verified By', 'Workflow - Accepted By', 'Workflow - Closed By']
        df_ccc[columns_to_preprocess] = df_ccc[columns_to_preprocess].apply(preprocess_column, mapping=mapping_dict)
        #st.write("Task-9")
        # Task -9: Related to Priority Name
        options = [
            "01. NOT Applicable",
            "02. Insufficient documents attached",
            "03. Incomplete Work",
            "03. Test Equipment not available",
            "04. Test Equipment not available",
            "05. No Power Supply available at site",
            "06. Material not as per data sheet/drawing",
            "07. Installation not as per drawing",
            "08. Unacceptable Test Procedure/Result",
            "09. No action for RFI on Hold",
            "10. Repeated/unnecessary RFI",
            "11. Test Crew not available",
            "12. Poor Workmanship",
            "13. PTW not available",
            "14. Spec. Violation",
            "15. Safety Issues",
            "16. Violation of ITP",
            "17. Others as indicated in Step 9 Comments:"
        ]

        # Update 'Comm Status' based on 'Inspection Result'
        df_ccc.loc[df_ccc['Inspection Result'] == 'Accepted', 'Comm Status'] = '01. NOT Applicable'
        df_ccc.loc[df_ccc['Inspection Result'] == 'Cancelled', 'Comm Status'] = '03. Incomplete Work'

        # rejected_count = df_ccc[df_ccc['Inspection Result'] == 'Rejected'].shape[0]
        # st.write(f"Number of rows with Inspection Result as 'Rejected': {rejected_count}")
        # If 'Inspection Result' is 'Rejected', show a dropdown list
        if (df_ccc['Inspection Result'] == 'Rejected').any():
            #st.write("Inside the IF")
            selected_option = st.selectbox('Select a reason:', options)
            df_ccc.loc[df_ccc['Inspection Result'] == 'Rejected', 'Comm Status'] = selected_option 
        # Drop 'Discipline' and 'Company Intervention Type:' columns
        
        df_ccc.drop(columns=['Discipline','Company Intervention Type:'], inplace=True)
        
        # Rename remaining columns
        df_ccc.columns=['Comm ID', 'Communication Type', 'Field011 (Custom)', 'Reply Date', 
                    'Workflow - Verified Date', 'Workflow - Verified By', 'Workflow - Accepted Date', 
                    'Workflow - Accepted By', 'Workflow - Closed Date', 'Workflow - Closed By','Task - (List Name)','Field012 (Custom)','Priority (Name)','Comm Status']
        df_ccc['Comments'] = None
        
        return df_ccc

    def get_binary_file_downloader_html(file_path, file_name):
        with open(file_path, 'rb') as f:
            data = f.read()
        bin_str = base64.b64encode(data).decode()
        href = f'<a href="data:application/octet-stream;base64,{bin_str}" download="{file_name}">Download</a>'
        return href

    # if __name__ == "__main__":


    st.write("**Programme Started .......**")
    # Boolean flag to control download option visibility
    show_download = False
    
    # Check if files are uploaded
    if uploaded_file_ccc is None or uploaded_file_name is None:
        st.write("Please upload CCC System File and Name Mapping File.")
    else:
        processed_df = preprocess_df_ccc(uploaded_file_ccc, uploaded_file_name)
        
        # Read the SCDB XLSX file as a DataFrame
        if uploaded_file_scdb is not None:
            df_SCDB = pd.read_excel(uploaded_file_scdb)
            
            if processed_df is not None and df_SCDB is not None:
                # Save the DataFrame to an Excel file
                output_file_path = 'processed_data.xlsx'
                processed_df.to_excel(output_file_path, index=False)

                # Load the workbook
                workbook = load_workbook(output_file_path)
                sheet = workbook.active

                columns_to_preprocess = ['Workflow - Verified By', 'Workflow - Accepted By', 'Workflow - Closed By']

                # Iterate over each column name
                for column_name in columns_to_preprocess:
                    # Get the column index from the column name
                    column_index = None
                    for col in sheet.iter_cols(min_row=1, max_row=1):
                        for cell in col:
                            if cell.value == column_name:
                                column_index = cell.column
                                break
                        if column_index:
                            break

                    if column_index:
                        for row in sheet.iter_rows(min_row=2, min_col=column_index, max_row=sheet.max_row, max_col=column_index):
                            for cell in row:
                                if cell.value == "No Name Match":
                                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    else:
                        print(f"Column '{column_name}' not found in the worksheet.")

                # Save the workbook
                workbook.save(output_file_path)
                workbook = None
                gc.collect()
                show_download = True
                # Provide a download button for the user to download the Excel file
                # st.markdown(get_binary_file_downloader_html(output_file_path, file_name='processed_data.xlsx'), unsafe_allow_html=True)
                # Add a button to trigger report generation
                if st.button("Generate report"):
                    # if show_download:
                    #     # Provide a download button for the user to download the Excel file
                    #     st.markdown(get_binary_file_downloader_html(output_file_path, file_name='processed_data.xlsx'), unsafe_allow_html=True)
                    #     st.write("***Programme Ended .......***")
                    # Extract unique Comm IDs from each DataFrame
                    processed_df_ids = set(processed_df['Comm ID'].unique())
                    df_SCDB_ids = set(df_SCDB['Comm ID'].unique())
                    # Find the mismatched IDs (IDs present in processed_df_ids but not in df_SCDB_ids)
                    mismatched_ids = processed_df_ids - df_SCDB_ids

                    if mismatched_ids:
                        st.write("Mismatched Comm IDs (IDs present in processed_df but not in df_SCDB):")
                        st.write(list(mismatched_ids)) 
                    else:
                        st.write("No mismatched Comm IDs found.")
                    # Missing Value Report
                    result = processed_df[['Workflow - Verified By', 'Workflow - Accepted Date', 
                                'Workflow - Accepted By', 'Workflow - Closed Date', 
                                'Workflow - Closed By','Priority (Name)','Field012 (Custom)']].isnull().sum()

                    st.write("***Null Value Counts for Each Column:***")
                    st.write(result)
                    # Missing Value Report per column with commID.

                    columns_to_check = ['Workflow - Verified By', 'Workflow - Accepted Date', 
                    'Workflow - Accepted By', 'Workflow - Closed Date', 
                    'Workflow - Closed By','Priority (Name)','Field012 (Custom)']

                    # Create an empty DataFrame to store the results
                    result_df = pd.DataFrame()

                    # Iterate over each column
                    for column in columns_to_check:
                        # Get the 'Comm ID' values where the column is null
                        comm_ids = processed_df.loc[processed_df[column].isnull(), 'Comm ID']
                        
                        # Create a DataFrame from the 'Comm ID' values
                        comm_ids_df = pd.DataFrame(comm_ids)
                        
                        # Add the DataFrame to the result DataFrame
                        result_df = pd.concat([result_df, comm_ids_df], axis=1)

                    # Set the column names of the result DataFrame
                    result_df.columns = columns_to_check

                    # Display the result DataFrame as a table
                    st.table(result_df)

                    missing_count = processed_df[processed_df['Priority (Name)'] != 'Cancelled']['Field012 (Custom)'].isnull().sum()
                    st.write("Count of missing values in 'Field012 (Custom)' where 'Priority (Name)' is not 'Cancelled':", missing_count)
                    if show_download:
                        # Provide a download button for the user to download the Excel file
                        st.markdown(get_binary_file_downloader_html(output_file_path, file_name='processed_data.xlsx'), unsafe_allow_html=True)
                        st.write("***Programme Ended .......***")
                
    






