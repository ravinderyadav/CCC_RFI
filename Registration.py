import streamlit as st
from PIL import Image
import base64
import pandas as pd
import os
from datetime import datetime
import base64
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import warnings
import time
warnings.filterwarnings('ignore')

def app():
    # List of image paths (replace with your own image paths)
    image_paths = [
        r"D:\Streamlit\img-4.png"
    ]

    # Create a placeholder for the image
    image_placeholder = st.empty()

    # Display each image in the list for 5 seconds
    for image_path in image_paths:
        # Load the image
        image = Image.open(image_path)
        
        # Display the image in the placeholder
        image_placeholder.image(image, use_column_width=True)
        
        # Wait for 5 seconds before displaying the next image
        time.sleep(0)



    def extract_date_from_filename(filename):
        # Extract the date portion from the file name
        date_str = filename.split('_')[-1].split('.')[0]  # Split by underscores and remove file extension
        return pd.to_datetime(date_str, format='%d-%m-%Y').strftime('%d-%b-%y')


    def preprocess_column(column, mapping_dict):
        """ Preprocess a column using a mapping dictionary for replacements """
        return column.map(lambda x: "".join(str(x).lower().split())).map(mapping_dict)

    def process_excel_file_1(uploaded_file, mapping_file):
        # Step 1: Reading and preprocessing data
        df = pd.read_excel(uploaded_file)

        # Get the file creation timestamp
        file_name = uploaded_file.name
        formatted_creation_date = extract_date_from_filename(file_name)

        # Subset and rename columns
        df = df[['Comm ID', 'Title', 'Communication Type', 'Responsible Company (Name)', 'Responsible Workgroup',
                'Asset Tag', 'Physical Location - Plant (Summary)', 'Physical Location - Facility (Summary)',
                'Physical Location Parent- Facility (Summary)', 'Workflow - Status', 'Document - Name/ID',
                'Discipline', 'Additional Information (Optional)', 'Inspection Coverage (Full/Partial):',
                'ITP No.', 'ITP Item No.', 'Inspection Time:', 'SubContractor Intervention Type:',
                'CTJV Intervention Type:', 'Company Intervention Type:', 'SC Construction PIC Name:',
                'SC QC PIC Name:', 'CTJV QC PIC Name:', 'Expected Duration (# of days):', 'Company PIC Name:',
                'Inspection Start Date:', 'Actual Inspection Finish Date:', 'Task Name', 'EXPECTED END DATE']]
        
        df.columns = ['Comm ID', 'Title', 'Communication Type', 'Responsible Company (Name)', 'Responsible Workgroup',
                    'Asset - (List Name)', 'Physical Location - Plant (Summary)', 'Physical Location - Location (Summary)',
                    'Physical Location Parent- Facility (Summary)', 'Workflow - Status', 'Document - (List Name)',
                    'Discipline', 'Information Requested', 'Disturbance Number', 'Field001 (Custom)',
                    'Field002 (Custom)', 'Field003 (Custom)', 'Field004 (Custom)', 'Field005 (Custom)',
                    'Field006 (Custom)', 'Field007 (Custom)', 'Field008 (Custom)', 'Field009 (Custom)',
                    'Field010 (Custom)', 'Field011 (Custom)', 'Due Date', 'Reply Date', 'Task - (List Name)',
                    'EXPECTED END DATE']
        
        # Mapping rules for 'Discipline' and 'Field009 (Custom)'
        discipline_mapping = {
            'AB': 'BLD', 'CVL': 'CVL', 'ELE': 'ELE', 'IN': 'CSE', 'IS': 'INS', 'MEC': 'MEC',
            'PA': 'PNT', 'PIP': 'PIP', 'STR': 'STR', 'TC': 'TEL', 'HV': 'MEC', 'RF': 'REF'
        }
        lead_mapping = {
            'BLD': 'Paglinawan, Cipriano/Gasang, Carla Joyzenia/Kanikodan, Prakashan',
            'CVL': 'Paglinawan, Cipriano/Gasang, Carla Joyzenia/Kanikodan, Prakashan',
            'ELE': 'Paglinawan, Cipriano/Gasang, Carla Joyzenia/Khurshid, Karawan',
            'CSE': 'Paglinawan, Cipriano/Gasang, Carla Joyzenia/Khurshid, Karawan',
            'TEL': 'Paglinawan, Cipriano/Gasang, Carla Joyzenia/Khurshid, Karawan',
            'PNT': 'Paglinawan, Cipriano/Gasang, Carla Joyzenia/Anjum, Mohammad Muzamil',
            'REF': 'Paglinawan, Cipriano/Gasang, Carla Joyzenia/Anjum, Mohammad Muzamil',
            'INS': 'Paglinawan, Cipriano/Gasang, Carla Joyzenia/Anjum, Mohammad Muzamil',
            'PIP': 'Paglinawan, Cipriano/Gasang, Carla Joyzenia/Ali, Ahsan',
            'STR': 'Paglinawan, Cipriano/Gasang, Carla Joyzenia/Agbay, Cecilio Masagpag',
            'MEC': 'Paglinawan, Cipriano/Gasang, Carla Joyzenia/Brosoto, Joelito'
        }
        
        df['Discipline'] = df['Discipline'].map(discipline_mapping)
        df['Field009 (Custom)'] = df['Discipline'].replace(lead_mapping)
        
        # Date format as per the requirement
        df[['EXPECTED END DATE', 'Due Date']] = df[['EXPECTED END DATE', 'Due Date']].apply(pd.to_datetime, errors='coerce')
        df['Date Difference'] = (df['EXPECTED END DATE'] - df['Due Date']).dt.days
        df[['EXPECTED END DATE', 'Due Date']] = df[['EXPECTED END DATE', 'Due Date']].apply(lambda x: x.dt.strftime('%d-%b-%y'))
        
        trigger_disciplines = ['MEC', 'PNT', 'PIP', 'STR', 'REF', 'INS']
        
        df['Information Requested'] = df.apply(lambda row:
            f"{row['Information Requested']} (INSPECTION DATE UPTO: {row['EXPECTED END DATE']})"
            if (row['Discipline'] in trigger_disciplines) & (row['Date Difference'] > 0)
            else row['Information Requested'], axis=1)
        
        df.drop(['EXPECTED END DATE', 'Date Difference'], axis=1, inplace=True)
        
        # Step 2: Preprocessing for Step-2
        mapping_df = pd.read_excel(mapping_file)
        mapping_dict = dict(zip(mapping_df.iloc[:, 0].str.lower().str.replace(' ', ''), mapping_df.iloc[:, 1]))
        
        step2 = df[['Comm ID', 'Communication Type']]
        step2['Workflow - Originated Date'] = formatted_creation_date
        step2['Workflow - Originated By'] = 'EL RAYES, Mohammad'
        step2['Workflow - Requested Date'] = formatted_creation_date
        step2['Workflow - Requested By'] = preprocess_column(df['Field008 (Custom)'], mapping_dict)
        step2['Workflow - Issued Date'] = datetime.now().strftime('%d-%b-%y')
        step2['Workflow - Issued By'] = 'Kunnummal, Sameer'
        
        step2.drop_duplicates(subset=['Comm ID'], keep='first', inplace=True)
        
        return df, step2

    def fill_missing_values_and_apply_conditions(sheet, columns_to_check):
        # Dictionary to keep track of row indices with missing values
        rows_with_missing_values = {}

        # Find the column indices based on column names
        column_indices = {column_name: sheet[1].index(column_name) + 1 for column_name in columns_to_check if column_name in sheet[1]}

        # Iterate over rows and mark rows with missing values
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            has_missing_value = any((row[col_idx - 1].value is None or row[col_idx - 1].value == "") for col_idx in column_indices.values())
            
            if has_missing_value:
                # Mark the row index as having missing values
                rows_with_missing_values[row_idx] = True

        # Apply row coloring based on rows with missing values
        for row_index in rows_with_missing_values:
            for cell in sheet[row_index]:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    def apply_conditions(sheet, all_columns):
        # Find indices of 'Discipline' and 'Comm ID' within all columns
        discipline_index = all_columns.index('Discipline') if 'Discipline' in all_columns else -1
        comm_id_index = all_columns.index('Comm ID') if 'Comm ID' in all_columns else -1

        if discipline_index != -1 and comm_id_index != -1:
            # Iterate over each row to apply conditions based on 'Comm ID'
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if len(row) > max(discipline_index, comm_id_index):
                    discipline_cell = row[discipline_index]
                    comm_id_cell = row[comm_id_index]

                    if discipline_cell.value and comm_id_cell.value:
                        discipline_value = str(discipline_cell.value).strip()
                        comm_id_value = str(comm_id_cell.value).strip()

                        if comm_id_value[9:11] != 'CL' and comm_id_value[9:11] != discipline_value[:2]:
                            # Apply fill color to the entire row
                            for cell in row:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        elif comm_id_value[9:11] == 'CL' and discipline_value[:2] != 'CV':
                            # Apply fill color to the entire row
                            for cell in row:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        else:
            st.warning("Column 'Discipline' or 'Comm ID' not found in the worksheet.")

    def process_file_2(uploaded_file,uploaded_file_SCDB):
        if uploaded_file is not None:
            # Load the workbook
            wb = load_workbook(uploaded_file)
            sheet = wb.active

            # Get all column names from the worksheet (first row)
            all_columns = [cell.value for cell in sheet[1] if cell.value]

            # Define the columns to check for missing values and apply conditions
            columns_to_check = [
                'Discipline',
                'Additional Information (Optional)',
                'Inspection Coverage (Full/Partial):',
                'ITP No.',
                'ITP Item No.',
                'Inspection Time:',
                'SubContractor Intervention Type:',
                'CTJV Intervention Type:',
                'Company Intervention Type:',
                'SC Construction PIC Name:',
                'SC QC PIC Name:',
                'Comm ID'
            ]

            # Process the worksheet
            fill_missing_values_and_apply_conditions(sheet, columns_to_check)
            apply_conditions(sheet, all_columns)

            # Extract base name of the uploaded file
            base_name, extension = os.path.splitext(uploaded_file.name)
            
            # Save the workbook with the base name of the uploaded file
            output_path = f"{base_name}{extension}"
            wb.save(output_path)
            st.success("Missing values filled and conditions applied.Output saved as " + output_path)

            # Download the output file
            st.download_button(
                label="Download Report File",
                data=open(output_path, "rb").read(),
                file_name=output_path,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # Close the workbook
            wb.close()
            # Summary
            df=pd.read_excel(uploaded_file)
            df2=pd.read_excel(uploaded_file_SCDB)
            df1 = set(df['Comm ID'].unique())
            df2 = set(df2['Comm ID'].unique())
            df3 = df1.intersection(df2)
            if df3:
                st.write("Common Comm ID's in both files:")
                st.write(list(df3)) 
            else:
                st.write("No common Comm ID's in both files.")
            # Show summary of missing values
            df=df[columns_to_check]
            st.subheader("Summary of Missing Values")
            # Identify columns with missing values
            columns_with_missing = df.isnull().sum().sum()

            if columns_with_missing:
                st.write("Somecolumns have missing values:")
            else:
                st.write("No columns have missing values.")
            missing_counts = df.isnull().sum()
            st.write(missing_counts)

    def main1():
        # apply Image
        # display_image()

        st.markdown("<h1 style='text-align: center; color: blue;'>CCC RFI Registration</h1>", unsafe_allow_html=True)
        
        # Section for choosing the program
        st.header("Choose Program")
        program_choice = st.selectbox("Choose the program to run:", ["Report","SCDB Data"])

        if program_choice == "SCDB Data":
            st.subheader("Upload the CCC System File and Name Mapping File")
            uploaded_data_file = st.file_uploader("CCC System File", type=["xlsx"])
            uploaded_mapping_file = st.file_uploader("Name Mapping File", type=["xlsx"])

            if st.button("Process Data for SCDB Data") and uploaded_data_file is not None and uploaded_mapping_file is not None:
                df_step1, df_step2 = process_excel_file_1(uploaded_data_file, uploaded_mapping_file)
                st.write("### File Processing Complete for SCDB Data!")

                # Download processed files
                st.write("### Download Processed Files for SCDB Data:")
                st.write("Download Step-1 CSV:")
                st.markdown(get_binary_file_downloader_html(df_step1, file_name="Step-1.csv"), unsafe_allow_html=True)
                st.write("Download Step-2 CSV:")
                st.markdown(get_binary_file_downloader_html(df_step2, file_name="Step-2.csv"), unsafe_allow_html=True)

        elif program_choice == "Report":
            st.subheader("Upload CCC File")
            uploaded_file_CCC = st.file_uploader("CCC Upload", type=["xlsx", "xls"])

            st.subheader("Upload SCDB")
            uploaded_file_SCDB = st.file_uploader("SCDB Upload", type=["xlsx", "xls"])

            if st.button("Process Files for Report") and uploaded_file_CCC is not None and uploaded_file_SCDB is not None:
                process_file_2(uploaded_file_CCC, uploaded_file_SCDB)  # Pass both uploaded files to the function


    def get_binary_file_downloader_html(df, file_name='file.csv'):
        """ Generates a link allowing the data in a given pandas dataframe to be downloaded """
        csv = df.to_csv(index=False)
        b64 = base64.b64encode(csv.encode()).decode()
        href = f'<a href="data:file/csv;base64,{b64}" download="{file_name}">Download CSV File</a>'
        return href


    main1()
